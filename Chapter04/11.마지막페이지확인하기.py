import requests
from bs4 import BeautifulSoup
import time
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 사용자 입력
keyword = pyautogui.prompt("검색어를 입력하세요")
lastpage = int(pyautogui.prompt("몇 페이지까지 크롤링 할까요?"))

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성하기
ws = wb.create_sheet(keyword)

# 열 너비 조절
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 120


# 행 번호
row = 1 

# 페이지 번호
page_num = 1
for i in range (1, lastpage * 10, 10) :
    print(f"{page_num} 페이지 크롤링 중입니다. ===============================")
    response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}&start={i}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    articles = soup.select("div.info_group") # 뉴스 기사 div 10개 추출

    for article in articles:
        links = article.select("a.info") # 리스트
        if len(links) >= 2 : # 링크가 2개 이상이면
            url = links[1].attrs['href'] # 두번째 링크의 href를 추출
            response = requests.get(url, headers={'User-agent':'Mozila/5.0'})
            html = response.text
            soup_sub = BeautifulSoup(html, 'html.parser')
    

            # 만약 연예 뉴스 라면
            if "entertain" in response.url :
                title = soup_sub.select_one(".end_tit")
                content = soup_sub.select_one("#articeBody")
            elif "sports" in response.url: 
                title = soup_sub.select_one("h4.title")
                content = soup_sub.select_one("#newsEndContents")
                # 본문 내용 안에 불필요한 div , p 삭제(compose)
                divs = content.select("div")     
                for div in divs :
                    div.decompose() 
                paragraphs = content.select("p")
                for p in paragraphs:
                    p.decompose()    
            else :
                title = soup_sub.select_one("#title_area")
                content = soup_sub.select_one("#contents")
            
            print("=========== 링크 ===========\n", url)
            print("=========== 제목 ===========\n", title.text.strip())
            print("=========== 본문 ===========\n", content.text.strip())
            ws[f'A{row}'] = url
            ws[f'B{row}'] = title.text.strip()
            ws[f'C{row}'] = content.text.strip()
            # 자동 줄 바꿈
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            row = row + 1 
            time.sleep(0.3) # 프로그램을 0.3초 잠시 멈추게 함

    # 마지막 페이지 여부 확인
    isLastPage = soup.select_one("a.btn_next").attrs['aria-disabled']
    if isLastPage == 'true':
        print("마지막 페이지 입니다")
        break
    page_num = page_num + 1

wb.save(f'{keyword}_result.xlsx')